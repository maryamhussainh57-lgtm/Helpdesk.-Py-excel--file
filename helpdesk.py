# helpdesk_excel_only.py
# Smart Student Help Desk System (Excel only)

from openpyxl import Workbook, load_workbook

print("Welcome to Smart Student Help Desk System!")
print("==========================================\n")

# Step 1: Get student details
student_name = input("Enter your name: ")

# Step 2: Select issue type
print("\nSelect your issue type:")
print("1. Academic Issue")
print("2. Fee Issue")
print("3. Technical Issue")

issue_type = input("Enter 1, 2, or 3: ")

# Step 3: Provide AI-style response offline
if issue_type == "1":
    response = "Check your timetable and contact your teacher."
    issue_desc = "Academic Issue"
elif issue_type == "2":
    response = "Check your fee slip and contact accounts department."
    issue_desc = "Fee Issue"
elif issue_type == "3":
    response = "Restart your device or check your internet connection."
    issue_desc = "Technical Issue"
else:
    response = "Invalid option. Please restart the program."
    issue_desc = "Invalid"

print("\nAI-Style Response:")
print(response)

# Step 4: Save record to Excel
file_name = "student_records.xlsx"
try:
    wb = load_workbook(file_name)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Name", "Issue Type", "Response"])  # header

sheet.append([student_name, issue_desc, response])
wb.save(file_name)
print(f"\nRecord saved to {file_name} ✅")

print("\nStudent gets guidance & solution. 🎉")